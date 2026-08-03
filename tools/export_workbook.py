"""
Dump everything the source workbook encodes into a machine-readable JSON file.

The Split Spreadsheet keeps its business logic in places that a normal "open
the file and look" pass misses: workbook-scoped LAMBDA definitions, array
formulas, data validation, conditional-format expressions and cell comments.
This script extracts all of them so the rules can be reviewed and diffed
rather than clicked through.

Usage:
    python tools/export_workbook.py reference/Expenses.xlsx docs/workbook-export.json
"""

from __future__ import annotations

import json
import re
import sys
import zipfile
from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

# Google Sheets exports formulas it cannot express in OOXML by wrapping them in
# IFERROR(__xludf.DUMMYFUNCTION("<the real formula>"), <last cached value>).
DUMMY = re.compile(
    r'^=IFERROR\(__xludf\.DUMMYFUNCTION\((.*)\),(.*)\)$', re.S
)


def unwrap(formula: str) -> tuple[str, str | None]:
    """Recovers the original Sheets formula and its last cached value."""
    match = DUMMY.match(formula)
    if not match:
        return formula, None

    literal, cached = match.groups()
    # The inner formula is an OOXML string literal, possibly split with & to
    # dodge the 255-character limit. Join the pieces and unescape the quotes.
    parts = re.findall(r'"((?:[^"]|"")*)"', literal)
    text = "".join(part.replace('""', '"') for part in parts)
    return text.replace("\\n", "\n"), cached.strip()


def cell_entries(worksheet, limit: int | None) -> list[dict]:
    entries = []
    max_row = worksheet.max_row if limit is None else min(limit, worksheet.max_row)

    for row in worksheet.iter_rows(min_row=1, max_row=max_row):
        for cell in row:
            if cell.value is None:
                continue

            entry: dict = {"ref": cell.coordinate}
            value = cell.value

            if isinstance(value, ArrayFormula):
                entry["kind"] = "array-formula"
                entry["spills"] = value.ref
                entry["formula"] = value.text
            elif isinstance(value, str) and value.startswith("="):
                formula, cached = unwrap(value)
                entry["kind"] = "formula"
                entry["formula"] = formula
                if cached is not None:
                    entry["exportedAs"] = value
                    entry["lastCachedValue"] = cached
            else:
                entry["kind"] = "value"
                entry["value"] = value

            entries.append(entry)

    return entries


def xml_fragments(archive: zipfile.ZipFile, part: str, tag: str) -> list[str]:
    if part not in archive.namelist():
        return []
    xml = archive.read(part).decode("utf-8")
    pattern = re.compile(rf"<{tag}\b.*?(?:/>|</{tag}>)", re.S)
    return [m.group(0) for m in pattern.finditer(xml)]


def comments(archive: zipfile.ZipFile) -> dict[str, dict[str, str]]:
    """Cell comments, keyed by comments part then cell reference."""
    out: dict[str, dict[str, str]] = {}
    for part in archive.namelist():
        if not re.match(r"xl/comments\d+\.xml", part):
            continue
        xml = archive.read(part).decode("utf-8")
        found = {}
        for match in re.finditer(r'<comment[^>]*ref="([^"]+)"[^>]*>(.*?)</comment>', xml, re.S):
            ref, body = match.groups()
            text = "".join(re.findall(r"<t[^>]*>(.*?)</t>", body, re.S))
            found[ref] = text.strip()
        out[part] = found
    return out


def export(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, data_only=False)
    archive = zipfile.ZipFile(path)

    sheet_parts = {
        sheet.title: f"xl/worksheets/sheet{index + 1}.xml"
        for index, sheet in enumerate(workbook.worksheets)
    }

    sheets = []
    for worksheet in workbook.worksheets:
        part = sheet_parts[worksheet.title]
        # The Currencies sheet is a 600-row static lookup table; a sample is
        # enough to document its shape without burying the interesting sheets.
        limit = 12 if worksheet.title == "Currencies" else None

        sheets.append(
            {
                "name": worksheet.title,
                "state": worksheet.sheet_state,
                "dimensions": worksheet.dimensions,
                "truncatedTo": limit,
                "cells": cell_entries(worksheet, limit),
                "dataValidations": xml_fragments(archive, part, "dataValidation"),
                "conditionalFormatting": xml_fragments(archive, part, "conditionalFormatting"),
            }
        )

    defined_names = {}
    for name, definition in workbook.defined_names.items():
        defined_names[name] = definition.value
    for worksheet in workbook.worksheets:
        for name, definition in worksheet.defined_names.items():
            defined_names[f"{worksheet.title}!{name}"] = definition.value

    return {
        "source": path.name,
        "hasVbaMacros": "xl/vbaProject.bin" in archive.namelist(),
        "parts": sorted(archive.namelist()),
        "definedNames": defined_names,
        "sheets": sheets,
        "comments": comments(archive),
    }


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 1

    source = Path(sys.argv[1])
    target = Path(sys.argv[2])
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        json.dumps(export(source), indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"wrote {target}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
